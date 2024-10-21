# app.py
from flask import Flask, request, jsonify, send_file, make_response
from flask_cors import CORS
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import json
import os
import logging
from datetime import datetime
import time
import random
from bs4 import BeautifulSoup
import requests
import re
from functools import wraps
from firebase_admin import credentials, auth as firebase_auth, initialize_app
import firebase_admin
from pymongo import MongoClient
from pymongo.server_api import ServerApi
from bson import ObjectId
from openpyxl import load_workbook
import glob
from io import BytesIO
from openpyxl import Workbook

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": ["https://aaaa-895ab.web.app", "http://localhost:3000", "http://localhost:5000", "https://8450-124-49-62-252.ngrok-free.app"]}}, supports_credentials=True)

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# MongoDB URI를 하드코딩
mongo_uri = "mongodb+srv://shinws8908:dnfhlao1@cluster0.h7c55.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0"
mongo_client = MongoClient(mongo_uri)
db = mongo_client['scraping_db']

# Firebase 서비스 계정 키 파일 경로를 하드코딩
service_account_key_path = 'serviceAccountKey.json'

# Firebase 초기화
cred = credentials.Certificate(service_account_key_path)
firebase_admin.initialize_app(cred)

def find_service_account_key():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    possible_names = ['serviceAccountKey.json', 'firebase-adminsdk.json', 'firebase-key.json']
    
    for name in possible_names:
        file_path = os.path.join(current_dir, name)
        if os.path.exists(file_path):
            return file_path
    
    # 파일을 찾지 못한 경우 glob을 사용하여 패턴 매칭
    for name in possible_names:
        pattern = os.path.join(current_dir, f'*{name}')
        matches = glob.glob(pattern)
        if matches:
            return matches[0]
    
    raise FileNotFoundError("서비스 계정 키 파일을 찾을 수 없습니.")

def require_uid(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        uid = request.args.get('uid')
        if not uid:
            return jsonify({"error": "UID가 제공되지 않았습니다."}), 400
        return f(uid, *args, **kwargs)
    return decorated_function

class NaverShoppingScraper:
    def __init__(self, uid):
        self.uid = uid
        self.user_data = db.users.find_one({"_id": uid})
        if not self.user_data:
            raise ValueError(f"User with UID {uid} not found")
        self.config = self.user_data.get('config', {})

    def setup_driver(self):
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--window-size=1200,800")
        chrome_options.add_argument("--incognito")
        chrome_options.add_experimental_option('detach', False)
        chrome_options.add_experimental_option("excludeSwitches", ["enable-logging", "enable-automation"])
        chrome_options.add_argument('log-level=3')
        chrome_options.add_argument(f"user-agent={self.config.get('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36')}")

        try:
            self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
            self.wait = WebDriverWait(self.driver, 30)
            logger.debug("WebDriver 설정 완료")
        except Exception as e:
            logger.error(f"WebDriver 정 중 오류 발생: {str(e)}")
            raise

    def should_skip_title(self, title):
        title_words = title.split()
        for word in title_words:
            if word in self.config.get('skip_words', []):
                return True
        if re.search(r"[^\w\s-]", title):
            return True
        return False

    def is_smartstore_product(self, product):
        mall_name = product.get('mallName', '').lower()
        is_valid = 'smartstore.naver.com' in product.get('mallProductUrl', '').lower()
        is_valid = is_valid and all(excluded.lower() not in mall_name for excluded in self.config.get('excluded_markets', []))
        logger.debug(f"Product ID {product.get('id', '')} is a smartstore product: {is_valid}")
        return is_valid

    def get_market_info(self, product):
        market_name = product.get('mallName', '')
        market_url = product.get('mallPcUrl', '')
        return market_name, market_url

    def download_json(self, keyword, sort_type):
        encoded_keyword = keyword.replace(" ", "%20")
        url = f"https://search.shopping.naver.com/search/all?adQuery={encoded_keyword}&origQuery={encoded_keyword}&pagingIndex=1&pagingSize=80&productSet=overseas&query={encoded_keyword}&sort={sort_type}"
        try:
            self.driver.get(url)
            logger.debug(f"URL 접근: {url}")
            time.sleep(random.uniform(2, 5))

            json_element = self.wait.until(EC.presence_of_element_located((By.ID, "__NEXT_DATA__")))
            json_data = json.loads(json_element.get_attribute('innerHTML'))
            logger.debug("JSON 데이터 로드 성공")
            return json_data
        except Exception as e:
            logger.error(f"JSON 데이터를 로드하는 중 오류 발생: {str(e)}")
            raise

    def process_products(self, keyword, products, sort_type_korean):
        processed_products = []
        for item in products:
            product = item.get('item', {})
            if self.is_smartstore_product(product):
                market_name, market_url = self.get_market_info(product)
                is_bundle = '묶음상품' if product.get('mallCount', 0) > 0 else '단품'
                related_keywords = product.get('relatedKeywords', [])
                related_keywords_str = ', '.join(related_keywords) if related_keywords else '없음'
                competitor_count = len(product.get('mallList', []))
                rating_score = product.get('scoreInfo', 0)
                search_trends = product.get('searchTrends', {}).get('trend', '정보 없음')
                rank = product.get('rank', None)

                row = {
                    "id": product.get('id', ''),
                    "keyword": keyword,
                    "market_name": market_name,
                    "product_title": product.get('productTitle', '상품명 없음'),
                    "price": product.get('price', '가격 없음'),
                    "product_url": product.get('mallProductUrl', ''),
                    "market_url": product.get('mallPcUrl', ''),
                    "manu_tag": product.get('manuTag', ''),
                    "image_url": product.get('imageUrl', ''),
                    "category": f"{product.get('category1Name', '')}>{product.get('category2Name', '')}>{product.get('category3Name', '')}",
                    "parentId": product.get('parentId', ''),
                    "category1Id": product.get('category1Id', ''),
                    "category2Id": product.get('category2Id', ''),
                    "category3Id": product.get('category3Id', ''),
                    "category4Id": product.get('category4Id', ''),
                    "brand": product.get('brand', ''),
                    "delivery_price": product.get('dlvryPrice', ''),
                    "is_naver_pay": product.get('isNaverPay', ''),
                    "review_count": product.get('reviewCount', 0),
                    "purchase_count": product.get('purchaseCnt', 0),
                    "recent_purchases": product.get('recentSales', 0),
                    "overseaTp": product.get('overseaTp', '1'),
                    "delivery_country": product.get('deliveryCountry', ''),
                    "product_type": product.get('type', ''),
                    "sort_type": sort_type_korean,
                    "is_bundle": is_bundle,
                    "related_keywords": related_keywords_str,
                    "search_trends": search_trends,
                    "competitor_count": competitor_count,
                    "rating_score": rating_score,
                    "rank": rank,
                    "mallInfoCache": product.get('mallInfoCache', {})
                }
                processed_products.append(row)
                logger.info(f"Added product data for keyword '{keyword}' with sort type '{sort_type_korean}'")
        return processed_products

    def search_products(self, keyword):
        try:
            self.setup_driver()
            all_products = []
            for sort_type in ['review', 'rel']:
                try:
                    json_data = self.download_json(keyword, sort_type)
                    products = json_data.get('props', {}).get('pageProps', {}).get('initialState', {}).get('products', {}).get('list', [])[:80]
                    sort_type_korean = '리뷰많은순' if sort_type == 'review' else '네이버랭킹순'
                    processed_products = self.process_products(keyword, products, sort_type_korean)
                    all_products.extend(processed_products)
                except Exception as e:
                    logger.error(f"{sort_type} 정렬로 검색 중 오류 발생: {str(e)}")
                    continue

            top_10_products = all_products[:10]
            logger.info(f"상위 10개 상품 추출 완료: {len(top_10_products)}개")

            # MongoDB에 검색 결과 저장
            db.users.update_one(
                {"_id": self.uid},
                {"$push": {"search_results": {"timestamp": datetime.now(), "products": top_10_products}}}
            )

            return top_10_products
        except Exception as e:
            logger.error(f"검색 중 오류 발생: {str(e)}")
            raise
        finally:
            if self.driver:
                self.driver.quit()
                logger.debug("WebDriver 료")

    def collect_selected_products(self, selected_product_ids):
        user_data = db.users.find_one({"_id": self.uid})
        search_results = user_data.get('search_results', [])
        
        if not search_results:
            return {"error": "검색 결과가 없습니다."}, 404

        latest_search_result = search_results[-1]['products']
        
        selected_products = [product for product in latest_search_result if str(product.get('id')) in selected_product_ids]
        
        if not selected_products:
            return {"error": "선택된 상품이 존재하지 않습니다."}, 404

        db.users.update_one(
            {"_id": self.uid},
            {"$push": {"collected_products": {"$each": selected_products}}}
        )

        return {
            "message": f"{len(selected_products)}개의 상품이 수집되었습니다.",
        }, 200

    def update_market_db(self, market_info):
        try:
            db.users.update_one(
                {"_id": self.uid},
                {"$set": {"market_db": market_info}}
            )
            logger.info(f"마켓 DB 업데이트 완료: {self.uid}")
            return {"message": "마켓 DB가 성공적으로 데트되었습니다."}
        except Exception as e:
            logger.error(f"마 DB 업데이트 중 오류 발생: {str(e)}")
            raise

def optimize_seo(title, keywords):
    # 실제 SEO 최적화 로직을 여기 구현
    # 임시로 간단 로을 사용
    optimized_title = f"{title} - {', '.join(keywords[:3])}"
    return optimized_title[:100]  # 제목 길이를 100자 제한

# 라우트 정의

@app.route('/')
def index():
    return "Welcome to AI Sourcing API", 200

@app.route('/health')
def health():
    return jsonify({"status": "healthy"}), 200

@app.route('/favicon.ico')
def favicon():
    return "", 204

@app.route('/signup', methods=['POST', 'OPTIONS'])
def signup():
    if request.method == 'OPTIONS':
        return '', 204
    
    logger.info("회원가입 요청 받음")
    data = request.get_json()
    logger.info(f"받은 데이터: {data}")
    try:
        email = data.get('email')
        password = data.get('password')
        display_name = data.get('displayName')

        if not email or not password or not display_name:
            logger.warning("필수 필드 누락")
            return jsonify({"error": "이메일, 비밀번호, 이름은 필수입니다."}), 400

        # Firebase로 사용자 생성
        user = firebase_auth.create_user(
            email=email,
            password=password,
            display_name=display_name
        )

        # MongoDB에 사용자 정보 저장
        user_data = {
            "_id": user.uid,
            "email": email,
            "display_name": display_name,
            "membershipLevel": "Basic",
            "remainingCredits": 10,
            "config": {
                "market": 0,
                "min_price": 0,
                "max_price": 1000000000,
                "option": "전체구매건수",
                "skip_words": [],
                "markets": []
            },
            "collected_products": [],
            "market_db": {},
            "search_results": []
        }
        db.users.insert_one(user_data)

        logger.info(f"회원가입 성공: {user.uid}")
        return jsonify({"success": True, "message": "User registered successfully", "uid": user.uid}), 200
    except Exception as e:
        logger.error(f"회원가입 오류: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 400

@app.route('/user-info', methods=['GET'])
@require_uid
def get_user_info(uid):
    try:
        user_data = db.users.find_one({"_id": uid})
        if user_data:
            return jsonify({
                "user": {
                    "name": user_data.get('display_name', '사용자'),
                    "email": user_data.get('email', ''),
                    "membershipLevel": user_data.get('membershipLevel', 'Basic'),
                    "remainingCredits": user_data.get('remainingCredits', 0)
                }
            }), 200
        else:
            return jsonify({"error": "사용자 정보를 찾을 수 없습니다."}), 404
    except Exception as e:
        logger.error(f"사용자 정보 조회 중 오류 발생: {str(e)}")
        return jsonify({"error": "사용자 정보 조회 중 오류가 발했습니다."}), 500

@app.route('/search', methods=['POST'])
@require_uid
def search_products_route(uid):
    try:
        data = request.get_json(force=True)
        keyword = data.get('keyword')
        if not keyword:
            logger.warning(f"키워드 없음: {uid}")
            return jsonify({"error": "키워드를 제공해야 합니다."}), 400
        
        logger.info(f"검색 요청 - 키워드: {keyword}, 사용자 ID: {uid}")
        
        scraper = NaverShoppingScraper(uid)
        products = scraper.search_products(keyword)
        
        logger.info(f"검색 완료 - 결과 수: {len(products)}")
        return jsonify({"products": products}), 200
    except Exception as e:
        logger.error(f"검색 중 오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"검색 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/collect', methods=['POST'])
@require_uid
def collect_product(uid):
    try:
        data = request.get_json()
        product_ids = data.get('selected_product_ids', [])
        
        # 여기에 상품 수집 로직 구현
        
        return jsonify({"message": f"{len(product_ids)}개의 상품이 수집되었습니다."}), 200
    except Exception as e:
        logger.error(f"상품 수집 중 오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"상품 수집 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/get_collected_products', methods=['GET'])
@require_uid
def get_collected_products(uid):
    try:
        user_data = db.users.find_one({"_id": uid})
        collected_products = user_data.get('collected_products', [])
        return jsonify({"products": collected_products}), 200
    except Exception as e:
        logger.error(f"수집된 상품 조회 중 오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"수집된 상품 조회 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/taobao_match', methods=['POST'])
@require_uid
def taobao_match(uid):
    try:
        data = request.get_json(force=True)
        image_url = data.get('image_url')
        product_id = data.get('productId')
        
        if not image_url:
            return jsonify({"error": "이미지 URL 제공되지 않았습니다."}), 400

        url = "https://open-taobao-api.p.rapidapi.com/taobao/traffic/item/imgsearch"
        querystring = {"pic_url": image_url, "language": "en"}
        headers = {
            "x-rapidapi-key": "5d86270c77mshb44bc07820b6f55p1e2c3ajsnfbfd2a9a7943",
            "x-rapidapi-host": "open-taobao-api.p.rapidapi.com"
        }

        response = requests.get(url, headers=headers, params=querystring)
        taobao_data = response.json()

        if isinstance(taobao_data, dict) and taobao_data.get('code') == 200 and taobao_data.get('data'):
            matched_item = taobao_data['data'][0]
            result = {
                "itemId": matched_item['itemId'],
                "title": matched_item.get('title') or matched_item.get('multiLanguageInfo', {}).get('title', ''),
                "price": matched_item['price'],
                "mainImageUrl": matched_item['mainImageUrl'],
                "shopName": matched_item['shopName'],
                "inventory": matched_item.get('inventory', 'N/A')
            }
            
            db.users.update_one(
                {"_id": uid, "collected_products.id": product_id},
                {"$set": {"collected_products.$.taobaoMatch": result}}
            )
            
            return jsonify(result), 200
        else:
            return jsonify({"error": "매칭된 상품이 없거나 API 응답이 올바르지 않습니다."}), 404

    except Exception as e:
        logger.error(f"타오바오 매칭  오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"타오바오 매칭  중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/seo_optimize', methods=['POST'])
@require_uid
def seo_optimize_route(uid):
    try:
        data = request.get_json(force=True)
        product_id = data.get('productId')
        
        user_data = db.users.find_one({"_id": uid})
        product = next((p for p in user_data.get('collected_products', []) if p['id'] == product_id), None)
        
        if not product:
            return jsonify({"error": "상품을 찾을 수 없습니다."}), 404
        
        related_keywords = product.get('related_keywords', '').split(',')
        optimized_title = optimize_seo(product['product_title'], related_keywords)
        
        db.users.update_one(
            {"_id": uid, "collected_products.id": product_id},
            {"$set": {"collected_products.$.seo_product_name": optimized_title}}
        )
        
        return jsonify({"optimized_title": optimized_title}), 200
    except Exception as e:
        logger.error(f"SEO 최적화 중 오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"SEO 최적화 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/get_market_db', methods=['GET'])
@require_uid
def get_market_db_route(uid):
    try:
        user_data = db.users.find_one({"_id": uid})
        if user_data and 'market_db' in user_data:
            return jsonify({"markets": list(user_data['market_db'].values())}), 200
        else:
            return jsonify({"markets": []}), 200
    except Exception as e:
        logger.error(f"마켓 DB 조회 중 오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"마켓 DB 조회 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/collect_market', methods=['POST'])
@require_uid
def collect_market_route(uid):
    try:
        data = request.get_json(force=True)
        market_data = data.get('market_data', [])

        if not market_data or not isinstance(market_data, list):
            return jsonify({"error": "유효한 마켓 데이터가 제공되지 않았습니다."}), 400

        scraper = NaverShoppingScraper(uid)
        result = scraper.update_market_db({m['mallName']: m for m in market_data})

        return jsonify(result), 200
    except Exception as e:
        logger.error(f"마켓 수집 요청 처리 중 오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"마켓 수집 요청 처리 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/calculate_shipping', methods=['POST'])
@require_uid
def calculate_shipping(uid):
    try:
        user_data = db.users.find_one({"_id": uid})
        collected_products = user_data.get('collected_products', [])
        
        for product in collected_products:
            if 'price_with_shipping' in product and 'price' in product:
                product['calculated_shipping_fee'] = product['price_with_shipping'] - product['price']
            else:
                product['calculated_shipping_fee'] = 0  # 또는 다른 기본값
        
        db.users.update_one({"_id": uid}, {"$set": {"collected_products": collected_products}})
        
        return jsonify({"message": "배송비 계산 완료"}), 200
    except Exception as e:
        logger.error(f"배송비 계산 중 오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"배송비 계산 중 오류가 발했습니다: {str(e)}"}), 500

@app.route('/download_heyseller', methods=['POST'])
@require_uid
def download_heyseller(uid):
    try:
        user_data = db.users.find_one({"_id": uid})
        collected_products = user_data.get('collected_products', [])
        
        template_data = db.templates.find_one({"name": "heyseller_template"})
        if not template_data:
            return jsonify({"error": "헤이셀러 템플릿을 찾을 수 없습니다."}), 404
        
        # 템플릿 데이터를 BytesIO 객체로 변환
        template_content = BytesIO(template_data['content'])
        
        try:
            wb = load_workbook(template_content)
        except Exception as e:
            # 템플릿 로드 실패 시 새 워북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "HeySeller Template"
            ws['A1'] = "카테고리"
            ws['B1'] = "상품명"
            ws['C1'] = "가격"
            ws['D1'] = "배송비"
            ws['E1'] = "타오바오 URL"
        
        ws = wb.active
        
        for idx, product in enumerate(collected_products, start=2):
            ws.cell(row=idx, column=1, value=product.get('category', ''))
            ws.cell(row=idx, column=2, value=product.get('seo_product_name', product.get('product_title', '')))
            ws.cell(row=idx, column=3, value=product.get('price', ''))
            ws.cell(row=idx, column=4, value=product.get('calculated_shipping_fee', ''))
            taobao_url = f"https://item.taobao.com/item.htm?id={product.get('taobaoMatch', {}).get('itemId', '')}"
            ws.cell(row=idx, column=5, value=taobao_url)
        
        # 메모리에 파일 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"헤이셀러_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"헤이셀러 파일 생성 중 오류 발생: {str(e)}", exc_info=True)
        return jsonify({"error": f"헤이셀러 파일 생성 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/login', methods=['POST', 'OPTIONS'])
def login():
    if request.method == 'OPTIONS':
        return '', 204
    app.logger.info("Login route accessed")
    data = request.get_json()
    app.logger.info(f"Login attempt for user: {data.get('email')}")
    
    try:
        # Firebase 토큰 검증
        id_token = request.headers.get('Authorization').split('Bearer ')[1]
        decoded_token = firebase_auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        
        # MongoDB에서 사용자 정보 확인
        user_data = db.users.find_one({"_id": uid})
        if not user_data:
            # 사용자 정보가 없으면 새로 생성
            user_data = {
                "_id": uid,
                "email": decoded_token.get('email'),
                "display_name": decoded_token.get('name'),
                "membershipLevel": "Basic",
                "remainingCredits": 10,
                "config": {
                    "market": 0,
                    "min_price": 0,
                    "max_price": 1000000000,
                    "option": "전체구매건수",
                    "skip_words": [],
                    "markets": []
                },
                "collected_products": [],
                "market_db": {},
                "search_results": []
            }
            db.users.insert_one(user_data)
        
        return jsonify({"success": True, "message": "Login successful", "uid": uid}), 200
    except Exception as e:
        app.logger.error(f"Login error: {str(e)}")
        return jsonify({"error": str(e)}), 400

@app.route('/google_login', methods=['POST', 'OPTIONS'])
def google_login():
    if request.method == 'OPTIONS':
        return '', 204
    data = request.get_json()
    # 여기에 Google 로그인 처리 로직 구현
    response = jsonify({"message": "Google 로그인 성공"})
    response.headers.add('Access-Control-Allow-Origin', 'http://127.0.0.1:3000')
    return response, 200

@app.errorhandler(Exception)
def handle_exception(e):
    app.logger.error(f"Unhandled exception: {str(e)}")
    return jsonify({"error": "An unexpected error occurred"}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": "Not found"}), 404

@app.errorhandler(405)
def method_not_allowed(error):
    return jsonify({"error": "Method not allowed"}), 405

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
